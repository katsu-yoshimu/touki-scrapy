import Message
import xlsContorller
import toukiController
from datetime import datetime

output_file_list = []

# Excel版設定ファイル読込
import openpyxl
from preflist import PREF_CODE

def getConfigFromXlsx():
    CONDITION_FILE_PATH = './Condition.xlsx'
    CONDITION_SHEET_NAME = '収集条件'

    # 返却値の初期化
    config = {
        "user_id"  : None,
        "password" : None,
        "conditions_list" : []
    }

    # デフォルト値の初期化
    todofukenShozai_def = ''
    chibanKuiki_def = ''
    chiban_from_def = ''
    chiban_to_def = ''

    # テンプレートxlsxを読込
    wb = openpyxl.load_workbook(CONDITION_FILE_PATH, data_only=True)

    # シートを指定
    ws = wb[CONDITION_SHEET_NAME]

    # ID番号
    config['user_id'] = ws.cell(row=2, column=4).value
    if config['user_id'] == None:
        config['user_id'] = ''

    # パスワード
    config['password'] = ws.cell(row=3, column=4).value
    if config['password'] == None:
        config['password'] = ''

    for i in range(6, 16):
        
        # 都道府県名
        todofukenShozai = ws.cell(row=i, column=3).value
        if todofukenShozai == None:
            todofukenShozai = todofukenShozai_def
        else:
            todofukenShozai = PREF_CODE[todofukenShozai]
            todofukenShozai_def = todofukenShozai

        # 市町村名
        chibanKuiki = ws.cell(row=i, column=4).value
        if chibanKuiki == None:
            chibanKuiki = chibanKuiki_def
        else:
            chibanKuiki_def = chibanKuiki
        
        # 地番・家屋番号 開始
        chiban_from = ws.cell(row=i, column=5).value
        if chiban_from == None:
            chiban_from = chiban_from_def
        else:
            chiban_from = str(chiban_from)
            chiban_from_def = chiban_from

        # 地番・家屋番号 終了
        chiban_to = ws.cell(row=i, column=6).value
        if chiban_to == None:
            chiban_to = chiban_to_def
        else:
            chiban_to = str(chiban_to)
            chiban_to_def = chiban_to

        # 請求種別
        seikyuJiko = ws.cell(row=i, column=7).value
        if seikyuJiko == '土地の全部事項':
            shozaiType = '土地'
            seikyuJiko = ['全部事項']

        elif seikyuJiko == '建物の全部事項':
            shozaiType = '建物'
            seikyuJiko = ['全部事項']

        elif seikyuJiko == '土地所在図/地積測量図':
            shozaiType = '土地'
            seikyuJiko = ['土地所在図/地積測量図']

        elif seikyuJiko == '建物図面/各階平面図':
            shozaiType = '建物'
            seikyuJiko = ['建物図面/各階平面図']
        else:
            seikyuJiko = []

        # 請求種別がある場合のみ処理対象
        if len(seikyuJiko) > 0:
            conditions = [
                shozaiType,
                todofukenShozai,
                chibanKuiki,
                chiban_from,
                chiban_to,
                seikyuJiko
            ]
            config['conditions_list'].append(conditions)
    return config

def main():
    # 設定ファイル読込
    config = None
    try:
        config = getConfigFromXlsx()
    except Exception as e:
        errorMessage=f'設定ファイルの読み込みでエラーが発生しました。\nエラー内容[{e}]'
        print(errorMessage)
        Message.MessageForefrontShowinfo(errorMessage)
        return
    
    user_id = config['user_id']
    password = config['password']
    conditions_list = config['conditions_list']

    # 実行時間確認
    if toukiController.is_RunEnable(datetime.now()) == False:
        errorMessage = f'現在は{datetime.now().strftime('%Y/%m/%d %H:%M:%S')}です。実行時間外です。\n平日8時30分～21時0分に実行してください。'
        print(errorMessage)
        Message.MessageForefrontShowinfo(errorMessage)
        return
    
    # 収集条件チェック(収集条件) 
    for conditions in conditions_list:
        if toukiController.checkConditions(conditions) == False:
            # エラーメッセージを表示し、処理終了
            errorMessage = '収集条件に誤りがあります。\nログを参照し、訂正後、再実行してください。'
            print(errorMessage)
            Message.MessageForefrontShowinfo(errorMessage)
            return

    # 収集開始メッセージ
    startMessage = f'不動産請求情報収集を実行しますか？\n実行ユーザはID番号【{user_id}】、パスワード【{password}】です。\n収集条件は以下の通りです。'
    for i, conditions in enumerate(conditions_list):
        startMessage += f'\n\n{i+1}：{xlsContorller.editCollectionCondition(conditions)}'
    print(startMessage)
    if Message.MessageForefront(startMessage) == False:
        return
    
    # データ収集
    for conditions in conditions_list:
        output_file_path = toukiController.collectData(conditions, user_id, password)
        # 処理終了時のメッセージ表示のため、出力ファイル名を追記
        output_file_list.append(output_file_path)
    
    # 収集終了メッセージ
    endMessage = '不動産請求情報収集が処理終了しました。\n収集結果は以下に出力されています。ご確認ください。'
    for i, output_file in enumerate(output_file_list):
        endMessage += f'\n{i+1}：【{output_file}】'
    print(endMessage)
    Message.MessageForefrontShowinfo(endMessage)
            

import os
import sys
# カレントディレクトリ変更
os.chdir(os.path.dirname(os.path.abspath(sys.argv[0])))

main()

# 終了時に自動的にコンソールを消さない
input("\n≪≪≪≪≪ コンソールを消すためには、「Enter」キーを押してください ≫≫≫≫≫\n")
