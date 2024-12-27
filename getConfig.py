# Excel版設定ファイル読込
import openpyxl
from preflist import PREF_CODE

def getConfigFromXlsx():
    CONDITION_FILE_PATH = './Condition.xlsm'
    CONDITION_SHEET_NAME = '収集条件'
    SETTING_SHEET_NAME = "設定値"

    # 返却値の初期化
    config = {
        "user_id"  : None,
        "password" : None,
        "conditions_list" : []
    }

    # デフォルト値の初期化
    todofukenShozai_def = ''
    chibanKuiki_def = ''
    chiban_from_def = ''
    chiban_to_def = ''

    # Condition.xlsmを読込
    wb = openpyxl.load_workbook(CONDITION_FILE_PATH, data_only=True)

    # シートを指定
    ws = wb[CONDITION_SHEET_NAME]

    # ID番号
    config['user_id'] = ws.cell(row=2, column=4).value
    if config['user_id'] == None:
        config['user_id'] = ''

    # パスワード
    config['password'] = ws.cell(row=3, column=4).value
    if config['password'] == None:
        config['password'] = ''

    for i in range(6, 16):
        
        # 都道府県名
        todofukenShozai = ws.cell(row=i, column=3).value
        if todofukenShozai == None:
            todofukenShozai = todofukenShozai_def
        else:
            todofukenShozai = PREF_CODE[todofukenShozai]
            todofukenShozai_def = todofukenShozai

        # 市町村名
        chibanKuiki = ws.cell(row=i, column=4).value
        if chibanKuiki == None:
            chibanKuiki = chibanKuiki_def
        else:
            chibanKuiki_def = chibanKuiki
        
        # 地番・家屋番号 開始
        chiban_from = ws.cell(row=i, column=5).value
        if chiban_from == None:
            chiban_from = chiban_from_def
        else:
            chiban_from = str(chiban_from)
            chiban_from_def = chiban_from

        # 地番・家屋番号 終了
        chiban_to = ws.cell(row=i, column=6).value
        if chiban_to == None:
            chiban_to = chiban_to_def
        else:
            chiban_to = str(chiban_to)
            chiban_to_def = chiban_to

        # 請求種別
        # チェックあり：True／なし：False
        # 7カラムから12カラムにチェックなしのときは無効行
        if ws.cell(row=i, column=7).value != True and \
           ws.cell(row=i, column=8).value != True and \
           ws.cell(row=i, column=9).value != True and \
           ws.cell(row=i, column=10).value != True and \
           ws.cell(row=i, column=11).value != True and \
           ws.cell(row=i, column=12).value != True:
            pass

        # 12カラムにチェックありのときは、地番・家屋番号の選択のみの動作を行う
        if ws.cell(row=i, column=12).value == True:    
            select_only = True

            # 7カラムから11カラムの一番左にチェックありに従う
            if ws.cell(row=i, column=7).value == True:
                shozaiType = '土地'
                seikyuJiko = ['全部事項']
            elif ws.cell(row=i, column=8).value == True:
                shozaiType = '土地'
                seikyuJiko = ['土地所在図/地積測量図']
            elif ws.cell(row=i, column=9).value == True:
                shozaiType = '土地'
                seikyuJiko = ['地役権図面']
            elif ws.cell(row=i, column=10).value == True:
                shozaiType = '建物'
                seikyuJiko = ['全部事項']
            elif ws.cell(row=i, column=11).value == True:
                shozaiType = '建物'
                seikyuJiko = ['建物図面/各階平面図']
            # 7カラムから11カラムのいずれもチェックなしのときは、7カラムがチェックされているものとして扱う
            else:
                shozaiType = '土地'
                seikyuJiko = ['全部事項']

            config['conditions_list'].append([shozaiType, todofukenShozai, chibanKuiki, chiban_from, chiban_to, seikyuJiko, select_only])

            # 12カラムにチェックありのときは、以降の行は無効行
            break

        # 12カラムにチェックなしのときは、7カラムから11カラムにチェックありのものを条件に追加
        else:
            select_only = False
            
            if ws.cell(row=i, column=7).value == True:
                shozaiType = '土地'
                seikyuJiko = ['全部事項']
                config['conditions_list'].append([shozaiType, todofukenShozai, chibanKuiki, chiban_from, chiban_to, seikyuJiko, select_only])
            if ws.cell(row=i, column=8).value == True:
                shozaiType = '土地'
                seikyuJiko = ['土地所在図/地積測量図']
                config['conditions_list'].append([shozaiType, todofukenShozai, chibanKuiki, chiban_from, chiban_to, seikyuJiko, select_only])
            if ws.cell(row=i, column=9).value == True:
                shozaiType = '土地'
                seikyuJiko = ['地役権図面']
                config['conditions_list'].append([shozaiType, todofukenShozai, chibanKuiki, chiban_from, chiban_to, seikyuJiko, select_only])
            if ws.cell(row=i, column=10).value == True:
                shozaiType = '建物'
                seikyuJiko = ['全部事項']
                config['conditions_list'].append([shozaiType, todofukenShozai, chibanKuiki, chiban_from, chiban_to, seikyuJiko, select_only])
            if ws.cell(row=i, column=11).value == True:
                shozaiType = '建物'
                seikyuJiko = ['建物図面/各階平面図']
                config['conditions_list'].append([shozaiType, todofukenShozai, chibanKuiki, chiban_from, chiban_to, seikyuJiko, select_only])

    # 設定値の読込＆設定
    ws = wb[SETTING_SHEET_NAME]
    config['MAX_CHIBAN_SELECT_NUMBER'] = ws.cell(row=3, column=5).value
    config['MAX_CHIBAN_INTERVAL']      = ws.cell(row=4, column=5).value
    config['MAX_WAIT_TIME']            = ws.cell(row=5, column=5).value
    config['CHIBAN_RETRY_WAIT_TIME']   = ws.cell(row=6, column=5).value
    config['CHIBAN_RETRY_OUT_COUNT']   = ws.cell(row=7, column=5).value
    config['INTERVAL_TIME']            = ws.cell(row=8, column=5).value
    config['INTERVAL_TIME_RATE']       = ws.cell(row=9, column=5).value
    config['IS_NOTIFY_WINDOWS']        = ws.cell(row=10, column=5).value
    config['IS_NOTIFY_BROWSER']        = ws.cell(row=11, column=5).value
    
    # 設定値の妥当性チェック
    if int != type(config["MAX_CHIBAN_SELECT_NUMBER"]):
        raise Exception("設定値：MAX_CHIBAN_SELECT_NUMBERの値は、1以上の整数を指定してください")
    elif config["MAX_CHIBAN_SELECT_NUMBER"] < 1:
        raise Exception("設定値：MAX_CHIBAN_SELECT_NUMBERの値は、1以上の整数を指定してください")
    
    if int != type(config["MAX_CHIBAN_INTERVAL"]):
        raise Exception("設定値：MAX_CHIBAN_INTERVALの値は、1以上の整数を指定してください")
    elif config["MAX_CHIBAN_INTERVAL"] < 1:
        raise Exception("設定値：MAX_CHIBAN_INTERVALの値は、1以上の整数を指定してください")
    
    if int != type(config["MAX_WAIT_TIME"]):
        raise Exception("設定値：MAX_WAIT_TIMEの値は、1以上の整数を指定してください")
    elif config["MAX_WAIT_TIME"] < 1:
        raise Exception("設定値：MAX_WAIT_TIMEの値は、1以上の整数を指定してください")
    
    if int != type(config["CHIBAN_RETRY_WAIT_TIME"]):
        raise Exception("設定値：CHIBAN_RETRY_WAIT_TIMEの値は、1以上の整数を指定してください")
    elif config["CHIBAN_RETRY_WAIT_TIME"] < 1:
        raise Exception("設定値：CHIBAN_RETRY_WAIT_TIMEの値は、1以上の整数を指定してください")

    if int != type(config["CHIBAN_RETRY_OUT_COUNT"]):
        raise Exception("設定値：CHIBAN_RETRY_OUT_COUNTの値は、1以上の整数を指定してください")
    elif config["CHIBAN_RETRY_OUT_COUNT"] < 1:
        raise Exception("設定値：CHIBAN_RETRY_OUT_COUNTの値は、1以上の整数を指定してください")

    if int != type(config["INTERVAL_TIME"]):
        raise Exception("設定値：INTERVAL_TIMEの値は、1以上の整数を指定してください")
    elif config["INTERVAL_TIME"] < 1:
        raise Exception("設定値：INTERVAL_TIMEの値は、1以上の整数を指定してください")
    
    if int != type(config["INTERVAL_TIME_RATE"]) and float != type(config["INTERVAL_TIME_RATE"]):
        raise Exception("設定値：INTERVAL_TIME_RATEの値は、0.0～1.0の小数を指定してください")
    elif config["INTERVAL_TIME_RATE"] < 0.0 or config["INTERVAL_TIME_RATE"] > 1.0:
        raise Exception("設定値：INTERVAL_TIME_RATEの値は、0.0～1.0の小数を指定してください")

    if int != type(config["IS_NOTIFY_WINDOWS"]):
        raise Exception("設定値：IS_NOTIFY_WINDOWSの値は、0：通知なし、または、1：通知ありを指定してください")
    elif config["IS_NOTIFY_WINDOWS"] != 0 and config["IS_NOTIFY_WINDOWS"] != 1:
        raise Exception("設定値：IS_NOTIFY_WINDOWSの値は、0：通知なし、または、1：通知ありを指定してください")
    
    if int != type(config["IS_NOTIFY_BROWSER"]):
        raise Exception("設定値：IS_NOTIFY_BROWSERの値は、0：通知なし、または、1：通知ありを指定してください")
    elif config["IS_NOTIFY_BROWSER"] != 0 and config["IS_NOTIFY_BROWSER"] != 1:
        raise Exception("設定値：IS_NOTIFY_BROWSERの値は、0：通知なし、または、1：通知ありを指定してください")
    
    wb.close()

    return config